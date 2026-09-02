import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

def generate_excel():
    drugs_data = [
        # The original 30
        ["Concor 5mg", "مزمن - ضغط", True, 85, 30],
        ["Amlopres 5mg", "مزمن - ضغط", True, 45, 30],
        ["Diovan 80mg", "مزمن - ضغط", True, 150, 30],
        ["Norvasc 5mg", "مزمن - ضغط", True, 60, 30],
        ["Capoten 25mg", "مزمن - ضغط", True, 35, 30],
        ["Glucophage 500mg", "مزمن - سكر", True, 40, 30],
        ["Glucomin 500mg", "مزمن - سكر", True, 35, 30],
        ["Galvus Met 50/1000mg", "مزمن - سكر", True, 220, 30],
        ["Amaryl 2mg", "مزمن - سكر", True, 90, 30],
        ["Diamicron MR 60mg", "مزمن - سكر", True, 110, 30],
        ["Lipitor 20mg", "مزمن - كوليسترول", True, 130, 30],
        ["Crestor 10mg", "مزمن - كوليسترول", True, 160, 30],
        ["Eltroxin 100mcg", "مزمن - غدة درقية", True, 25, 30],
        ["Panadol Extra", "مسكنات", False, 20, 10],
        ["Bi Alcofan", "مسكنات", False, 30, 10],
        ["Cataflam 50mg", "مسكنات", False, 25, 10],
        ["Brufen 400mg", "مسكنات", False, 22, 10],
        ["Amoclan 1g", "مضاد حيوي", False, 65, 7],
        ["Augmentin 1g", "مضاد حيوي", False, 90, 7],
        ["Zinnat 500mg", "مضاد حيوي", False, 110, 7],
        ["Rani 150mg", "جهاز هضمي", False, 15, 15],
        ["Motilium 10mg", "جهاز هضمي", False, 30, 15],
        ["Nexium 40mg", "جهاز هضمي", False, 95, 20],
        ["Cal-D-Vita", "فيتامينات", False, 65, 30],
        ["Bio Vit-C 1000mg", "فيتامينات", False, 55, 30],
        ["Feroglobin", "فيتامينات", False, 120, 30],
        ["Omega Life 3", "فيتامينات", False, 140, 30],
        ["Coldrex", "نزلات برد", False, 25, 10],
        ["Comtrex", "نزلات برد", False, 28, 10],
        ["Zyrtec 10mg", "حساسية", False, 45, 20],
        
        # Adding 20 more common drugs as a bonus
        ["Telfast 120mg", "حساسية", False, 60, 20],
        ["Claritine 10mg", "حساسية", False, 40, 20],
        ["Levhistam 5mg", "حساسية", False, 35, 20],
        ["Congestal", "نزلات برد", False, 20, 10],
        ["1,2,3 Tablets", "نزلات برد", False, 15, 10],
        ["Flumox 500mg", "مضاد حيوي", False, 45, 7],
        ["Hibiotic 1g", "مضاد حيوي", False, 75, 7],
        ["Ketofan 50mg", "مسكنات", False, 15, 10],
        ["Voltaren 50mg", "مسكنات", False, 35, 10],
        ["Panadol Advance", "مسكنات", False, 20, 10],
        ["Controloc 40mg", "جهاز هضمي", False, 85, 20],
        ["Antinal", "جهاز هضمي", False, 25, 10],
        ["Spasmo-Digestin", "جهاز هضمي", False, 22, 15],
        ["Centrum", "فيتامينات", False, 150, 30],
        ["Osteocare", "فيتامينات", False, 70, 30],
        ["Neuroton", "فيتامينات", False, 45, 30],
        ["Janumet 50/1000mg", "مزمن - سكر", True, 250, 30],
        ["Novomix 30 Penfill", "مزمن - سكر", True, 180, 30],
        ["Ator 20mg", "مزمن - كوليسترول", True, 65, 30],
        ["L-Thyroxin 50mcg", "مزمن - غدة درقية", True, 30, 30]
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Egyptian Drugs Seed"
    
    # Headers
    headers = ["name", "category", "is_chronic", "base_price", "default_cycle_days", "image_url"]
    ws.append(headers)
    
    # Style headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Add data
    for d in drugs_data:
        row = d + [""] # Empty string for image_url
        ws.append(row)
        
    # Set column widths
    ws.column_dimensions['A'].width = 25 # name
    ws.column_dimensions['B'].width = 20 # category
    ws.column_dimensions['C'].width = 12 # is_chronic
    ws.column_dimensions['D'].width = 12 # base_price
    ws.column_dimensions['E'].width = 18 # default_cycle_days
    ws.column_dimensions['F'].width = 40 # image_url
    
    # Right-to-Left alignment for Arabic category column
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    save_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "seed_data", "Egyptian_Drugs_Seed.xlsx"))
    wb.save(save_path)
    print(f"Excel sheet created successfully at: {save_path}")

if __name__ == "__main__":
    generate_excel()
