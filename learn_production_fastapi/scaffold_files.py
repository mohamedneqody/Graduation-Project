import os

base_path = r"d:\Graduation Project\learn_production_fastapi"

files = {
    "app/routers/files.py": '''"""
# مسؤوليته:
# إدارة مسارات (Endpoints) رفع وتنزيل الملفات (Files).
# يحتوي على الـ Routers التي تتعامل مع UploadFile والـ Responses بأنواعها.
# 
# ممنوع أن يحتوي على:
# - المنطق المعقد لمعالجة الملفات (يتم نقله للـ services إن كبر حجمه).
"""
import uuid
import os
import csv
from io import StringIO
from fastapi import APIRouter, UploadFile, File, HTTPException, status
from fastapi.responses import FileResponse, StreamingResponse
from app.services import files as file_service

router = APIRouter(prefix="/api/v1/files", tags=["Files"])

@router.post("/upload/image")
async def upload_image(file: UploadFile = File(...)):
    # التحقق من نوع الملف (Content-Type)
    if file.content_type not in ["image/jpeg", "image/png", "image/jpg"]:
        raise HTTPException(status_code=400, detail="Invalid image type. Only JPG/PNG are allowed.")
    
    # حفظ الملف باستخدام الـ Service
    filename = await file_service.save_file(file, "images")
    return {"message": "Image uploaded successfully", "filename": filename}

@router.post("/upload/pdf")
async def upload_pdf(file: UploadFile = File(...)):
    # التحقق من النوع
    if file.content_type != "application/pdf":
        raise HTTPException(status_code=400, detail="Invalid file type. Only PDF is allowed.")
    
    # التحقق من حجم الملف قبل المعالجة والحفظ الكامل (أقل من 5MB)
    MAX_SIZE = 5 * 1024 * 1024 # 5MB
    content = await file.read()
    if len(content) > MAX_SIZE:
        raise HTTPException(status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE, detail="File too large. Max size is 5MB.")
    
    # إرجاع المؤشر لبداية الملف لأننا قرأناه
    await file.seek(0)
    
    filename = await file_service.save_file(file, "pdfs")
    return {"message": "PDF uploaded successfully", "filename": filename}

@router.post("/upload/excel")
async def upload_excel(file: UploadFile = File(...)):
    if not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed.")
    
    # قراءة الملف بدون حفظه
    result = await file_service.process_excel(file)
    return result

@router.post("/upload/csv")
async def upload_csv(file: UploadFile = File(...)):
    if file.content_type not in ["text/csv", "application/vnd.ms-excel"]:
        raise HTTPException(status_code=400, detail="Only CSV files are allowed.")
    
    content = await file.read()
    decoded_content = content.decode("utf-8")
    
    csv_reader = csv.reader(StringIO(decoded_content))
    rows = list(csv_reader)
    
    row_count = len(rows)
    col_count = len(rows[0]) if row_count > 0 else 0
    
    return {"row_count": row_count, "column_count": col_count}

@router.get("/download/{filename}")
async def download_file(filename: str):
    # البحث عن الملف في كل المجلدات كمثال بسيط
    file_path = file_service.get_file_path(filename)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(path=file_path, filename=filename, media_type="application/octet-stream")

@router.get("/stream/{filename}")
async def stream_file(filename: str):
    file_path = file_service.get_file_path(filename)
    if not file_path:
        raise HTTPException(status_code=404, detail="File not found")
    
    return StreamingResponse(file_service.file_iterator(file_path), media_type="application/octet-stream")
''',

    "app/services/files.py": '''"""
# مسؤوليته:
# معالجة الملفات، حفظها، والتعامل مع المكتبات الخارجية للقراءة مثل openpyxl.
"""
import os
import uuid
import aiofiles
import openpyxl
from io import BytesIO
from fastapi import UploadFile

UPLOAD_DIR = "uploads"

async def save_file(file: UploadFile, subfolder: str) -> str:
    """
    يحفظ الملف باسم فريد (UUID) لتجنب هجمات Path Traversal واختراق الملفات.
    """
    os.makedirs(os.path.join(UPLOAD_DIR, subfolder), exist_ok=True)
    
    # استخراج امتداد الملف الأصلي
    extension = os.path.splitext(file.filename)[1]
    
    # إنشاء اسم جديد تماماً
    unique_filename = f"{uuid.uuid4()}{extension}"
    file_path = os.path.join(UPLOAD_DIR, subfolder, unique_filename)
    
    # استخدام aiofiles للكتابة بشكل Async
    async with aiofiles.open(file_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
        
    return unique_filename

def get_file_path(filename: str) -> str | None:
    """يبحث عن الملف في مجلدات الرفع ويرجع المسار إن وجد"""
    for root, dirs, files in os.walk(UPLOAD_DIR):
        if filename in files:
            return os.path.join(root, filename)
    return None

async def file_iterator(file_path: str):
    """
    يقرأ الملف في أجزاء (Chunks) لتوفير الذاكرة أثناء الـ Streaming
    """
    async with aiofiles.open(file_path, mode="rb") as f:
        while chunk := await f.read(1024 * 1024):  # قراءة 1 ميجا في كل مرة
            yield chunk

async def process_excel(file: UploadFile) -> dict:
    content = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    
    return {
        "row_count": sheet.max_row,
        "column_count": sheet.max_column
    }
'''
}

for rel_path, content in files.items():
    full_path = os.path.join(base_path, rel_path.replace("/", os.sep))
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "w", encoding="utf-8") as f:
        f.write(content)

print("Files generated successfully.")
