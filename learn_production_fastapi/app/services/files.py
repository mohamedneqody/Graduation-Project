"""
# مسؤوليته:
# معالجة الملفات، حفظها، والتعامل مع المكتبات الخارجية للقراءة مثل openpyxl.
"""
import os
import uuid
import aiofiles
import openpyxl
from io import BytesIO
from fastapi import UploadFile

UPLOAD_DIR = "uploads"

async def save_file(file: UploadFile, subfolder: str) -> str:
    """
    يحفظ الملف باسم فريد (UUID) لتجنب هجمات Path Traversal واختراق الملفات.
    """
    os.makedirs(os.path.join(UPLOAD_DIR, subfolder), exist_ok=True)
    
    # استخراج امتداد الملف الأصلي
    extension = os.path.splitext(file.filename)[1]
    
    # إنشاء اسم جديد تماماً
    unique_filename = f"{uuid.uuid4()}{extension}"
    file_path = os.path.join(UPLOAD_DIR, subfolder, unique_filename)
    
    # استخدام aiofiles للكتابة بشكل Async
    async with aiofiles.open(file_path, 'wb') as out_file:
        content = await file.read()
        await out_file.write(content)
        
    return unique_filename

def get_file_path(filename: str) -> str | None:
    """يبحث عن الملف في مجلدات الرفع ويرجع المسار إن وجد"""
    for root, dirs, files in os.walk(UPLOAD_DIR):
        if filename in files:
            return os.path.join(root, filename)
    return None

async def file_iterator(file_path: str):
    """
    يقرأ الملف في أجزاء (Chunks) لتوفير الذاكرة أثناء الـ Streaming
    """
    async with aiofiles.open(file_path, mode="rb") as f:
        while chunk := await f.read(1024 * 1024):  # قراءة 1 ميجا في كل مرة
            yield chunk

async def process_excel(file: UploadFile) -> dict:
    content = await file.read()
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    sheet = workbook.active
    
    return {
        "row_count": sheet.max_row,
        "column_count": sheet.max_column
    }
